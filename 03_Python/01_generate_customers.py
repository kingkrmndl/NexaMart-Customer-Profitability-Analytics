from pathlib import Path
from datetime import date
import random
import pandas as pd
from faker import Faker
from openpyxl import load_workbook


# ============================================================
# 1. PROJECT PATHS
# ============================================================

PROJECT_DIR = Path(__file__).resolve().parent.parent
RAW_DATA_DIR = PROJECT_DIR / "01_Raw_Data"

EXCEL_FILE = RAW_DATA_DIR / "NexaMart_Raw_Data.xlsx"


# ============================================================
# 2. INITIALIZE TOOLS
# ============================================================

fake = Faker("en_IN")

random.seed(42)
Faker.seed(42)


# ============================================================
# 3. CUSTOMER DESIGN
# ============================================================

NUM_CUSTOMERS = 1000

SEGMENTS = (
    ["Consumer"] * 600
    + ["Small Business"] * 250
    + ["Corporate"] * 150
)


LOCATION_DATA = {
    "West Bengal": {
        "region": "East",
        "cities": ["Kolkata", "Siliguri", "Durgapur"]
    },
    "Maharashtra": {
        "region": "West",
        "cities": ["Mumbai", "Pune", "Nagpur"]
    },
    "Karnataka": {
        "region": "South",
        "cities": ["Bengaluru", "Mysuru", "Mangaluru"]
    },
    "Tamil Nadu": {
        "region": "South",
        "cities": ["Chennai", "Coimbatore", "Madurai"]
    },
    "Delhi": {
        "region": "North",
        "cities": ["New Delhi", "Delhi"]
    },
    "Telangana": {
        "region": "South",
        "cities": ["Hyderabad", "Warangal"]
    },
    "Gujarat": {
        "region": "West",
        "cities": ["Ahmedabad", "Surat", "Vadodara"]
    },
    "Uttar Pradesh": {
        "region": "North",
        "cities": ["Lucknow", "Noida", "Kanpur"]
    }
}


# ============================================================
# 4. GENERATE CUSTOMER RECORDS
# ============================================================

customers = []

states = list(LOCATION_DATA.keys())

for i in range(1, NUM_CUSTOMERS + 1):

    customer_id = f"CUS{i:04d}"

    customer_name = fake.name()

    customer_segment = SEGMENTS[i - 1]

    state = random.choice(states)

    city = random.choice(
        LOCATION_DATA[state]["cities"]
    )

    customer_since = fake.date_between(
    start_date=date(2021, 1, 1),
    end_date=date(2025, 12, 31)
)

    customers.append({
        "Customer_ID": customer_id,
        "Customer_Name": customer_name,
        "Customer_Segment": customer_segment,
        "City": city,
        "State": state,
        "Customer_Since": customer_since
    })


# ============================================================
# 5. CONVERT TO DATAFRAME
# ============================================================

customers_df = pd.DataFrame(customers)


# ============================================================
# 6. VALIDATION
# ============================================================

print("Number of customers:", len(customers_df))

print(
    "Unique Customer IDs:",
    customers_df["Customer_ID"].nunique()
)

print("\nCustomer segment distribution:")
print(
    customers_df["Customer_Segment"].value_counts()
)

print("\nFirst five customers:")
print(
    customers_df.head()
)


# ============================================================
# 7. WRITE DATA TO EXCEL
# ============================================================

workbook = load_workbook(EXCEL_FILE)

worksheet = workbook["Customers"]

# Remove anything below the header
if worksheet.max_row > 1:
    worksheet.delete_rows(
        2,
        worksheet.max_row - 1
    )

for row_number, row in enumerate(
    customers_df.itertuples(index=False),
    start=2
):

    for column_number, value in enumerate(
        row,
        start=1
    ):

        worksheet.cell(
            row=row_number,
            column=column_number,
            value=value
        )


workbook.save(EXCEL_FILE)


# ============================================================
# 8. FINAL MESSAGE
# ============================================================

print(
    "\nCustomer data successfully written to:"
)

print(EXCEL_FILE)