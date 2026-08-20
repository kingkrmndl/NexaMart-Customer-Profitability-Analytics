from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


# ============================================================
# 1. PROJECT PATHS
# ============================================================

PROJECT_DIR = Path(__file__).resolve().parent.parent

RAW_DATA_DIR = PROJECT_DIR / "01_Raw_Data"

EXCEL_FILE = RAW_DATA_DIR / "NexaMart_Raw_Data.xlsx"


# ============================================================
# 2. START
# ============================================================

print("=" * 60)

print("NexaMart — Write Validated Orders to Excel")

print("=" * 60)


print("\nExcel file:")

print(EXCEL_FILE)


if not EXCEL_FILE.exists():

    raise FileNotFoundError(
        f"Excel file not found: {EXCEL_FILE}"
    )


# ============================================================
# 3. IMPORT VALIDATED ORDER GENERATION
# ============================================================

print("\nLoading validated order-generation engine...")


from importlib.util import (
    spec_from_file_location,
    module_from_spec
)


GENERATOR_FILE = (

    PROJECT_DIR

    / "03_Python"

    / "02_generate_orders.py"
)


spec = spec_from_file_location(
    "order_generator",
    GENERATOR_FILE
)


order_generator = module_from_spec(
    spec
)


spec.loader.exec_module(
    order_generator
)


# ============================================================
# 4. GET VALIDATED ORDERS
# ============================================================

orders_df = (

    order_generator

    .GENERATED_ORDERS

    .copy()
)


print(
    "\nValidated orders received:",
    len(orders_df)
)


# ============================================================
# 5. FINAL SAFETY CHECK
# ============================================================

if len(orders_df) != 10000:

    raise ValueError(
        "Expected 10,000 orders."
    )


if (

    orders_df[
        "Order_ID"
    ]

    .nunique()

    !=

    len(orders_df)

):

    raise ValueError(
        "Duplicate Order IDs detected."
    )


if orders_df.isna().sum().sum() > 0:

    raise ValueError(
        "Missing values detected."
    )


print(
    "Final safety checks passed."
)


# ============================================================
# 6. DEFINE EXCEL COLUMNS
# ============================================================

excel_columns = [

    "Order_ID",

    "Order_Date",

    "Customer_ID",

    "Product_ID",

    "Region",

    "Sales_Channel",

    "Quantity",

    "Unit_Price",

    "Discount",

    "Unit_Cost",

    "Gross_Sales",

    "Discount_Amount",

    "Revenue",

    "COGS",

    "Gross_Profit",

    "Gross_Margin"
]


# ============================================================
# 7. CHECK COLUMNS
# ============================================================

missing_columns = (

    set(excel_columns)

    -

    set(orders_df.columns)
)


if missing_columns:

    raise ValueError(
        f"Missing columns: {missing_columns}"
    )


orders_to_write = (

    orders_df[
        excel_columns
    ]

    .copy()
)


# ============================================================
# 8. BACKUP ORIGINAL WORKBOOK
# ============================================================

backup_file = (

    RAW_DATA_DIR

    / "NexaMart_Raw_Data_Backup.xlsx"
)


print(
    "\nCreating workbook backup..."
)


if backup_file.exists():

    print(
        "Backup already exists:"
    )

    print(
        backup_file
    )

else:

    import shutil

    shutil.copy2(
        EXCEL_FILE,
        backup_file
    )

    print(
        "Backup created:"
    )

    print(
        backup_file
    )


# ============================================================
# 9. LOAD WORKBOOK
# ============================================================

print(
    "\nOpening Excel workbook..."
)


workbook = load_workbook(
    EXCEL_FILE
)


# ============================================================
# 10. CHECK ORDERS SHEET
# ============================================================

if "Orders" not in workbook.sheetnames:

    workbook.close()

    raise ValueError(
        "Orders worksheet not found."
    )


orders_sheet = workbook[
    "Orders"
]


# ============================================================
# 11. CLEAR EXISTING ORDERS DATA
# ============================================================

print(
    "\nExisting Orders rows:",
    orders_sheet.max_row
)


print(
    "Clearing existing Orders sheet..."
)


# Keep the header row.
# Delete everything below row 1.

if orders_sheet.max_row > 1:

    orders_sheet.delete_rows(
        2,
        orders_sheet.max_row - 1
    )


# ============================================================
# 12. WRITE HEADERS
# ============================================================

for column_number, column_name in enumerate(

    excel_columns,

    start=1

):

    orders_sheet.cell(

        row=1,

        column=column_number

    ).value = column_name


# ============================================================
# 13. WRITE DATA
# ============================================================

print(
    "\nWriting 10,000 validated orders..."
)


for row_number, row in enumerate(

    orders_to_write.itertuples(
        index=False,
        name=None
    ),

    start=2

):

    for column_number, value in enumerate(

        row,

        start=1

    ):

        # Convert pandas Timestamp
        # into a Python datetime.

        if isinstance(
            value,
            pd.Timestamp
        ):

            value = value.to_pydatetime()


        # Convert NumPy numeric values
        # into standard Python values.

        elif hasattr(
            value,
            "item"
        ):

            value = value.item()


        orders_sheet.cell(

            row=row_number,

            column=column_number

        ).value = value


# ============================================================
# 14. FORMAT HEADER
# ============================================================

for cell in orders_sheet[1]:

    cell.font = cell.font.copy(
        bold=True
    )


# ============================================================
# 15. SAVE WORKBOOK
# ============================================================

print(
    "\nSaving workbook..."
)


workbook.save(
    EXCEL_FILE
)


workbook.close()


# ============================================================
# 16. REOPEN AND VERIFY
# ============================================================

print(
    "Workbook saved successfully."
)


print(
    "\nReopening workbook for verification..."
)


verification_df = pd.read_excel(

    EXCEL_FILE,

    sheet_name="Orders"
)


# ============================================================
# 17. VERIFY ROW COUNT
# ============================================================

print(
    "\nOrders rows in Excel:",
    len(verification_df)
)


if len(verification_df) != 10000:

    raise ValueError(
        "Excel row count is not 10,000."
    )


# ============================================================
# 18. VERIFY ORDER IDs
# ============================================================

unique_ids = (

    verification_df[
        "Order_ID"
    ]

    .nunique()
)


print(
    "Unique Order IDs in Excel:",
    unique_ids
)


if unique_ids != 10000:

    raise ValueError(
        "Excel contains duplicate Order IDs."
    )


# ============================================================
# 19. VERIFY MISSING VALUES
# ============================================================

missing_values = (

    verification_df

    .isna()

    .sum()

    .sum()
)


print(
    "Missing values in Excel:",
    missing_values
)


if missing_values > 0:

    raise ValueError(
        "Excel contains missing values."
    )


# ============================================================
# 20. VERIFY FINANCIAL TOTALS
# ============================================================

excel_revenue = (

    verification_df[
        "Revenue"
    ]

    .sum()
)


excel_profit = (

    verification_df[
        "Gross_Profit"
    ]

    .sum()
)


generated_revenue = (

    orders_df[
        "Revenue"
    ]

    .sum()
)


generated_profit = (

    orders_df[
        "Gross_Profit"
    ]

    .sum()
)


print(
    "\nGenerated Revenue:",
    round(
        generated_revenue,
        2
    )
)


print(
    "Excel Revenue:",
    round(
        excel_revenue,
        2
    )
)


print(
    "\nGenerated Gross Profit:",
    round(
        generated_profit,
        2
    )
)


print(
    "Excel Gross Profit:",
    round(
        excel_profit,
        2
    )
)


# ============================================================
# 21. FINANCIAL RECONCILIATION
# ============================================================

revenue_difference = abs(

    generated_revenue

    -

    excel_revenue
)


profit_difference = abs(

    generated_profit

    -

    excel_profit
)


print(
    "\nRevenue difference:",
    revenue_difference
)


print(
    "Profit difference:",
    profit_difference
)


if revenue_difference > 0.01:

    raise ValueError(
        "Revenue reconciliation failed."
    )


if profit_difference > 0.01:

    raise ValueError(
        "Profit reconciliation failed."
    )


# ============================================================
# 22. FINAL SUCCESS
# ============================================================

print("\n" + "=" * 60)

print(
    "ORDERS WRITE COMPLETE"
)

print("=" * 60)


print(
    "\n10,000 validated orders successfully written to:"
)


print(
    EXCEL_FILE
)


print(
    "\nBackup available at:"
)


print(
    backup_file
)


print(
    "\nRevenue reconciliation: PASSED"
)


print(
    "Gross Profit reconciliation: PASSED"
)


print(
    "Row count verification: PASSED"
)


print(
    "Order ID verification: PASSED"
)


print(
    "Missing-value verification: PASSED"
)


print(
    "\nThe Orders sheet is now populated."
)