import pandas as pd
from pathlib import Path


# ============================================================
# NexaMart Calendar Generation
# ============================================================

print("=" * 60)
print("NexaMart Calendar Generation")
print("=" * 60)


# ------------------------------------------------------------
# 1. PROJECT PATHS
# ------------------------------------------------------------

PROJECT_ROOT = Path(r"G:\Customer_Profitability_Analytics")

RAW_DATA_FOLDER = PROJECT_ROOT / "01_Raw_Data"

EXCEL_FILE = RAW_DATA_FOLDER / "NexaMart_Raw_Data.xlsx"


print("\nExcel file:")
print(EXCEL_FILE)

print("\nExcel file exists:", EXCEL_FILE.exists())


if not EXCEL_FILE.exists():
    raise FileNotFoundError(
        f"Excel file not found:\n{EXCEL_FILE}"
    )


# ------------------------------------------------------------
# 2. LOAD ORDERS
# ------------------------------------------------------------

print("\n" + "=" * 60)
print("LOADING ORDERS")
print("=" * 60)


orders = pd.read_excel(
    EXCEL_FILE,
    sheet_name="Orders"
)


print(f"\nOrders loaded: {len(orders):,}")


# ------------------------------------------------------------
# 3. CHECK ORDER DATE
# ------------------------------------------------------------

print("\n" + "=" * 60)
print("CHECKING ORDER DATE")
print("=" * 60)


if "Order_Date" not in orders.columns:
    raise ValueError(
        "Order_Date column not found in Orders sheet."
    )


orders["Order_Date"] = pd.to_datetime(
    orders["Order_Date"],
    errors="coerce"
)


invalid_dates = orders["Order_Date"].isna().sum()


print(f"\nInvalid order dates: {invalid_dates}")


if invalid_dates > 0:
    raise ValueError(
        "Some Order_Date values could not be converted to dates."
    )


min_date = orders["Order_Date"].min().normalize()
max_date = orders["Order_Date"].max().normalize()


print(f"Minimum order date: {min_date}")
print(f"Maximum order date: {max_date}")


# ------------------------------------------------------------
# 4. GENERATE COMPLETE DATE RANGE
# ------------------------------------------------------------

print("\n" + "=" * 60)
print("GENERATING CALENDAR")
print("=" * 60)


calendar_dates = pd.date_range(
    start=min_date,
    end=max_date,
    freq="D"
)


calendar = pd.DataFrame({
    "Calendar_Date": calendar_dates
})


print(
    f"\nCalendar dates generated: {len(calendar):,}"
)


# ------------------------------------------------------------
# 5. CREATE CALENDAR ATTRIBUTES
# ------------------------------------------------------------

calendar["Year"] = calendar["Calendar_Date"].dt.year

calendar["Quarter"] = (
    "Q"
    + calendar["Calendar_Date"].dt.quarter.astype(str)
)

calendar["Month"] = calendar["Calendar_Date"].dt.month

calendar["Month_Name"] = (
    calendar["Calendar_Date"]
    .dt.month_name()
)


# Optional useful analytical fields

calendar["Year_Month"] = (
    calendar["Calendar_Date"]
    .dt.to_period("M")
    .astype(str)
)

calendar["Month_Start"] = (
    calendar["Calendar_Date"]
    .dt.to_period("M")
    .dt.start_time
)

calendar["Month_End"] = (
    calendar["Calendar_Date"]
    .dt.to_period("M")
    .dt.end_time
    .dt.normalize()
)

calendar["Week_Number"] = (
    calendar["Calendar_Date"]
    .dt.isocalendar()
    .week
    .astype(int)
)

calendar["Day_Name"] = (
    calendar["Calendar_Date"]
    .dt.day_name()
)

calendar["Day_of_Week"] = (
    calendar["Calendar_Date"]
    .dt.dayofweek + 1
)


# ------------------------------------------------------------
# 6. VALIDATION
# ------------------------------------------------------------

print("\n" + "=" * 60)
print("CALENDAR VALIDATION")
print("=" * 60)


print(
    f"\nCalendar rows: {len(calendar):,}"
)

print(
    f"Calendar columns: {len(calendar.columns)}"
)


print("\nColumns:")

for column in calendar.columns:
    print(f" - {column}")


# Check duplicate dates

duplicate_dates = (
    calendar["Calendar_Date"]
    .duplicated()
    .sum()
)


print(
    f"\nDuplicate calendar dates: {duplicate_dates}"
)


# Check missing values

missing_values = (
    calendar.isna()
    .sum()
    .sum()
)


print(
    f"Total missing values: {missing_values}"
)


if duplicate_dates != 0:
    raise ValueError(
        "Calendar contains duplicate dates."
    )


if missing_values != 0:
    raise ValueError(
        "Calendar contains missing values."
    )


# ------------------------------------------------------------
# 7. DISPLAY SAMPLE
# ------------------------------------------------------------

print("\n" + "=" * 60)
print("FIRST 10 CALENDAR ROWS")
print("=" * 60)


print(
    calendar.head(10).to_string(index=False)
)


print("\n" + "=" * 60)
print("LAST 10 CALENDAR ROWS")
print("=" * 60)


print(
    calendar.tail(10).to_string(index=False)
)


# ------------------------------------------------------------
# ------------------------------------------------------------
# 8. WRITE CALENDAR BACK TO EXCEL
# ------------------------------------------------------------

print("\n" + "=" * 60)
print("WRITING CALENDAR TO EXCEL")
print("=" * 60)


from openpyxl import load_workbook


# Open the existing workbook without rebuilding
# the other worksheets.

workbook = load_workbook(
    EXCEL_FILE
)


# Check that Calendar sheet exists.

if "Calendar" not in workbook.sheetnames:

    workbook.close()

    raise ValueError(
        "Calendar worksheet not found."
    )


calendar_sheet = workbook["Calendar"]


print(
    "\nExisting Calendar rows:",
    calendar_sheet.max_row
)


# ------------------------------------------------------------
# Clear existing Calendar data
# ------------------------------------------------------------

print(
    "Clearing existing Calendar sheet..."
)


if calendar_sheet.max_row > 1:

    calendar_sheet.delete_rows(
        2,
        calendar_sheet.max_row - 1
    )


# ------------------------------------------------------------
# Write headers
# ------------------------------------------------------------

for column_number, column_name in enumerate(
    calendar.columns,
    start=1
):

    calendar_sheet.cell(
        row=1,
        column=column_number
    ).value = column_name


# ------------------------------------------------------------
# Write Calendar data
# ------------------------------------------------------------

print(
    "\nWriting calendar data..."
)


for row_number, row in enumerate(
    calendar.itertuples(
        index=False,
        name=None
    ),
    start=2
):

    for column_number, value in enumerate(
        row,
        start=1
    ):

        # Convert pandas Timestamp
        # to Python datetime.

        if isinstance(
            value,
            pd.Timestamp
        ):

            value = value.to_pydatetime()


        # Convert NumPy values
        # to standard Python values.

        elif hasattr(
            value,
            "item"
        ):

            value = value.item()


        calendar_sheet.cell(
            row=row_number,
            column=column_number
        ).value = value


# ------------------------------------------------------------
# Format header
# ------------------------------------------------------------

for cell in calendar_sheet[1]:

    cell.font = cell.font.copy(
        bold=True
    )


# ------------------------------------------------------------
# Save workbook
# ------------------------------------------------------------

print(
    "\nSaving workbook..."
)


workbook.save(
    EXCEL_FILE
)


workbook.close()


print(
    "\nCalendar sheet successfully written."
)

# ------------------------------------------------------------
# ------------------------------------------------------------
# 9. FINAL CHECK
# ------------------------------------------------------------

print("\n" + "=" * 60)
print("FINAL CHECK")
print("=" * 60)


verification = pd.read_excel(
    EXCEL_FILE,
    sheet_name="Calendar"
)


print(
    f"\nCalendar rows in Excel: {len(verification):,}"
)


print(
    f"Calendar columns in Excel: {len(verification.columns)}"
)


print(
    f"Minimum calendar date: "
    f"{verification['Calendar_Date'].min()}"
)


print(
    f"Maximum calendar date: "
    f"{verification['Calendar_Date'].max()}"
)


# ------------------------------------------------------------
# Final validation
# ------------------------------------------------------------

if len(verification) != len(calendar):

    raise ValueError(
        "Calendar row count changed after writing."
    )


if (
    verification["Calendar_Date"]
    .duplicated()
    .sum()
    != 0
):

    raise ValueError(
        "Duplicate Calendar_Date values found in Excel."
    )


if (
    verification.isna()
    .sum()
    .sum()
    != 0
):

    raise ValueError(
        "Missing values found in Calendar sheet."
    )


print(
    "\nCalendar verification passed."
)


print("\n" + "=" * 60)
print("CALENDAR GENERATION COMPLETE")
print("=" * 60)


print(
    "\nThe Calendar sheet is now ready."
)


print(
    "\nNext step: run 05_data_quality_audit.py"
)