from pathlib import Path
import random

import pandas as pd
from openpyxl import load_workbook


# ============================================================
# 1. PROJECT PATHS
# ============================================================

PROJECT_DIR = Path(__file__).resolve().parent.parent

RAW_DATA_DIR = PROJECT_DIR / "01_Raw_Data"

EXCEL_FILE = RAW_DATA_DIR / "NexaMart_Raw_Data.xlsx"


# ============================================================
# 2. RANDOMNESS
# ============================================================

random.seed(42)


# ============================================================
# 3. ORDER SETTINGS
# ============================================================

NUM_ORDERS = 10000

ORDER_START_DATE = pd.Timestamp("2023-01-01")

ORDER_END_DATE = pd.Timestamp("2025-12-31")


# ============================================================
# 4. START
# ============================================================

print("=" * 60)

print("NexaMart Order Generation")

print("=" * 60)


print("\nExcel file:")

print(EXCEL_FILE)


print(
    "\nExcel file exists:",
    EXCEL_FILE.exists()
)


if not EXCEL_FILE.exists():

    raise FileNotFoundError(
        f"Excel file not found: {EXCEL_FILE}"
    )


# ============================================================
# 5. CHECK WORKBOOK
# ============================================================

workbook = load_workbook(
    EXCEL_FILE,
    read_only=True,
    data_only=True
)


print("\nWorksheets found:")

print(workbook.sheetnames)


required_sheets = [
    "Orders",
    "Customers",
    "Products",
    "Calendar"
]


for sheet in required_sheets:

    if sheet not in workbook.sheetnames:

        workbook.close()

        raise ValueError(
            f"Required sheet missing: {sheet}"
        )


workbook.close()


# ============================================================
# 6. LOAD CUSTOMERS
# ============================================================

customers_df = pd.read_excel(
    EXCEL_FILE,
    sheet_name="Customers"
)


# ============================================================
# 7. LOAD PRODUCTS
# ============================================================

products_df = pd.read_excel(
    EXCEL_FILE,
    sheet_name="Products"
)


print(
    "\nCustomer records loaded:",
    len(customers_df)
)


print(
    "Unique Customer IDs:",
    customers_df["Customer_ID"].nunique()
)


print(
    "\nProduct records loaded:",
    len(products_df)
)


print(
    "Unique Product IDs:",
    products_df["Product_ID"].nunique()
)


# ============================================================
# 8. REQUIRED CUSTOMER COLUMNS
# ============================================================

required_customer_columns = [

    "Customer_ID",

    "Customer_Name",

    "Customer_Segment",

    "City",

    "State",

    "Customer_Since"
]


for column in required_customer_columns:

    if column not in customers_df.columns:

        raise ValueError(
            f"Missing customer column: {column}"
        )


# ============================================================
# 9. REQUIRED PRODUCT COLUMNS
# ============================================================

required_product_columns = [

    "Product_ID",

    "Product_Name",

    "Category",

    "Subcategory",

    "Brand"
]


for column in required_product_columns:

    if column not in products_df.columns:

        raise ValueError(
            f"Missing product column: {column}"
        )


# ============================================================
# 10. PRODUCT PRICING RULES
# ============================================================

PRICING_RULES = {

    # --------------------------------------------------------
    # Electronics
    # --------------------------------------------------------

    "Wireless Mouse": (900, 0.58),

    "Mechanical Keyboard": (2800, 0.60),

    "USB-C Hub": (1800, 0.57),

    "Laptop Stand": (1500, 0.55),

    "Webcam": (2500, 0.62),

    "Bluetooth Speaker": (2200, 0.58),

    "Wireless Earbuds": (3500, 0.60),

    "Noise Cancelling Headphones": (7500, 0.62),

    "Portable Speaker": (4500, 0.59),

    "Soundbar": (9000, 0.64),


    # --------------------------------------------------------
    # Furniture
    # --------------------------------------------------------

    "Ergonomic Office Chair": (7500, 0.65),

    "Executive Office Chair": (12000, 0.68),

    "Computer Desk": (8000, 0.67),

    "Standing Desk": (15000, 0.69),

    "Folding Table": (4500, 0.64),

    "Bookshelf": (6000, 0.66),

    "Filing Cabinet": (7000, 0.70),

    "Storage Rack": (5500, 0.67),

    "Drawer Unit": (4000, 0.65),

    "Office Storage Box": (1200, 0.60),


    # --------------------------------------------------------
    # Office Supplies
    # --------------------------------------------------------

    "Ballpoint Pen Pack": (250, 0.48),

    "Gel Pen Pack": (300, 0.49),

    "Notebook": (180, 0.52),

    "Spiral Notebook": (250, 0.51),

    "Copy Paper Ream": (450, 0.62),

    "Desk Organizer": (500, 0.55),

    "Stapler": (300, 0.52),

    "Staple Pack": (120, 0.50),

    "Scissors": (180, 0.54),

    "Calculator": (650, 0.60),


    # --------------------------------------------------------
    # Home & Lifestyle
    # --------------------------------------------------------

    "Electric Kettle": (1800, 0.62),

    "Coffee Maker": (4500, 0.65),

    "Air Fryer": (6500, 0.67),

    "Water Bottle": (700, 0.55),

    "Lunch Box": (600, 0.54),

    "Table Lamp": (1200, 0.56),

    "Wall Clock": (900, 0.58),

    "LED Light Strip": (1000, 0.55),

    "Air Purifier": (12000, 0.70),

    "Humidifier": (5000, 0.66),


    # --------------------------------------------------------
    # Sports & Fitness
    # --------------------------------------------------------

    "Yoga Mat": (900, 0.55),

    "Dumbbell Set": (2500, 0.62),

    "Resistance Bands": (800, 0.53),

    "Fitness Tracker": (4000, 0.61),

    "Exercise Bike": (25000, 0.72),

    "Running Shoes": (3500, 0.64),

    "Football": (900, 0.57),

    "Cricket Bat": (3000, 0.63),

    "Badminton Racket": (1800, 0.59),

    "Sports Backpack": (1500, 0.56)
}


# ============================================================
# 11. CHECK PRICING COVERAGE
# ============================================================

missing_prices = (

    set(products_df["Product_Name"])

    -

    set(PRICING_RULES.keys())
)


print(
    "\nProducts without pricing rules:",
    len(missing_prices)
)


if missing_prices:

    print("\nMissing products:")

    print(
        sorted(missing_prices)
    )

    raise ValueError(
        "Some products do not have pricing rules."
    )


else:

    print(
        "All products have pricing rules."
    )


# ============================================================
# 12. CUSTOMER BEHAVIOUR FUNCTION
# ============================================================

def get_customer_behavior(segment):

    # --------------------------------------------------------
    # Consumer
    # --------------------------------------------------------

    if segment == "Consumer":

        quantity = random.randint(
            1,
            3
        )

        discount = random.choice(
            [
                0.00,
                0.05,
                0.10
            ]
        )

        sales_channel = random.choices(

            ["Online", "Store"],

            weights=[
                70,
                30
            ],

            k=1
        )[0]


    # --------------------------------------------------------
    # Small Business
    # --------------------------------------------------------

    elif segment == "Small Business":

        quantity = random.randint(
            1,
            8
        )

        discount = random.choice(
            [
                0.00,
                0.05,
                0.10,
                0.15
            ]
        )

        sales_channel = random.choices(

            ["Online", "Store"],

            weights=[
                55,
                45
            ],

            k=1
        )[0]


    # --------------------------------------------------------
    # Corporate
    # --------------------------------------------------------

    elif segment == "Corporate":

        quantity = random.randint(
            2,
            15
        )

        discount = random.choice(
            [
                0.05,
                0.10,
                0.15,
                0.20
            ]
        )

        sales_channel = random.choices(

            ["Online", "Store"],

            weights=[
                35,
                65
            ],

            k=1
        )[0]


    else:

        raise ValueError(
            f"Unknown customer segment: {segment}"
        )


    return (
        quantity,
        discount,
        sales_channel
    )


# ============================================================
# 13. STATE → REGION MAPPING
# ============================================================

STATE_TO_REGION = {

    "West Bengal": "East",

    "Maharashtra": "West",

    "Karnataka": "South",

    "Tamil Nadu": "South",

    "Delhi": "North",

    "Telangana": "South",

    "Gujarat": "West",

    "Uttar Pradesh": "North"
}


# ============================================================
# 14. VALIDATE STATES
# ============================================================

customer_states = sorted(

    customers_df["State"]

    .dropna()

    .unique()
)


print(
    "\nStates found in customer data:"
)


for state in customer_states:

    print(state)


unmapped_states = (

    set(customer_states)

    -

    set(STATE_TO_REGION.keys())
)


print(
    "\nNumber of unique states:",
    len(customer_states)
)


print(
    "Unmapped states:",
    len(unmapped_states)
)


if unmapped_states:

    print(
        "\nUnmapped states:"
    )

    print(
        sorted(unmapped_states)
    )

    raise ValueError(
        "Unmapped customer states found."
    )


else:

    print(
        "All customer states have region mappings."
    )


# ============================================================
# 15. CUSTOMER ORDER WEIGHTS
# ============================================================

CUSTOMER_SEGMENT_WEIGHTS = {

    "Consumer": 1,

    "Small Business": 2,

    "Corporate": 4
}


customers_df["Order_Weight"] = (

    customers_df["Customer_Segment"]

    .map(
        CUSTOMER_SEGMENT_WEIGHTS
    )
)


if customers_df["Order_Weight"].isna().any():

    raise ValueError(
        "Some customers have no order weight."
    )


print(
    "\nOrder weights by customer segment:"
)


print(

    customers_df

    .groupby(
        "Customer_Segment"
    )["Order_Weight"]

    .agg(
        [
            "count",
            "min",
            "max"
        ]
    )
)


# ============================================================
# 16. PREPARE CUSTOMER SAMPLING
# ============================================================

customer_choices = (

    customers_df[
        "Customer_ID"
    ]

    .tolist()
)


customer_weights = (

    customers_df[
        "Order_Weight"
    ]

    .tolist()
)

# ============================================================
# 17. GENERATE ORDERS
# ============================================================

orders = []


for i in range(
    1,
    NUM_ORDERS + 1
):

    # --------------------------------------------------------
    # Select customer
    # --------------------------------------------------------

    selected_customer_id = random.choices(
        customer_choices,
        weights=customer_weights,
        k=1
    )[0]


    # --------------------------------------------------------
    # Retrieve customer
    # --------------------------------------------------------

    customer = (
        customers_df[
            customers_df["Customer_ID"] == selected_customer_id
        ]
        .iloc[0]
    )


    # --------------------------------------------------------
    # Customer information
    # --------------------------------------------------------

    customer_segment = customer[
        "Customer_Segment"
    ]


    customer_state = customer[
        "State"
    ]


    # --------------------------------------------------------
    # Region
    # --------------------------------------------------------

    region = STATE_TO_REGION[
        customer_state
    ]


    # --------------------------------------------------------
    # Customer behaviour
    # --------------------------------------------------------

    (
        quantity,
        discount,
        sales_channel
    ) = get_customer_behavior(
        customer_segment
    )


    # --------------------------------------------------------
    # Select product
    # --------------------------------------------------------

    product = (
        products_df
        .sample(
            n=1
        )
        .iloc[0]
    )


    # --------------------------------------------------------
    # Product information
    # --------------------------------------------------------

    product_id = product[
        "Product_ID"
    ]


    product_name = product[
        "Product_Name"
    ]


    # --------------------------------------------------------
    # Pricing rule
    # --------------------------------------------------------

    (
        base_price,
        cost_ratio
    ) = PRICING_RULES[
        product_name
    ]


    # --------------------------------------------------------
    # Price variation
    # --------------------------------------------------------

    price_variation = random.uniform(
        0.95,
        1.05
    )


    unit_price = round(
        base_price * price_variation,
        2
    )


    # --------------------------------------------------------
    # Unit cost
    # --------------------------------------------------------

    unit_cost = round(
        unit_price * cost_ratio,
        2
    )


    # --------------------------------------------------------
    # Customer registration date
    # --------------------------------------------------------

    customer_since = pd.to_datetime(
        customer["Customer_Since"]
    )


    # --------------------------------------------------------
    # Earliest valid order date
    # --------------------------------------------------------
    #
    # An order cannot happen before the customer existed.
    #
    # Example:
    #
    # Customer_Since = 2024-06-15
    # Order_Start_Date = 2023-01-01
    #
    # Earliest valid order date = 2024-06-15
    #
    # --------------------------------------------------------

    earliest_order_date = max(
        ORDER_START_DATE,
        customer_since
    )


    # --------------------------------------------------------
    # Validate customer eligibility
    # --------------------------------------------------------

    if earliest_order_date > ORDER_END_DATE:

        raise ValueError(
            f"Customer {selected_customer_id} has "
            f"Customer_Since={customer_since.date()} "
            f"which is after the order generation end date "
            f"{ORDER_END_DATE.date()}."
        )


    # --------------------------------------------------------
    # Generate random order date
    # --------------------------------------------------------

    available_days = (
        ORDER_END_DATE
        - earliest_order_date
    ).days


    random_days = random.randint(
        0,
        available_days
    )


    order_date = (
        earliest_order_date
        + pd.Timedelta(
            days=random_days
        )
    )


    # --------------------------------------------------------
    # Final order-date validation
    # --------------------------------------------------------

    if order_date < customer_since:

        raise ValueError(
            f"Invalid Order_Date generated for "
            f"Customer {selected_customer_id}: "
            f"Order_Date={order_date.date()}, "
            f"Customer_Since={customer_since.date()}."
        )


    if order_date < ORDER_START_DATE:

        raise ValueError(
            f"Order date {order_date.date()} "
            f"is before the project start date "
            f"{ORDER_START_DATE.date()}."
        )


    if order_date > ORDER_END_DATE:

        raise ValueError(
            f"Order date {order_date.date()} "
            f"is after the project end date "
            f"{ORDER_END_DATE.date()}."
        )


    # --------------------------------------------------------
    # Append order
    # --------------------------------------------------------

    orders.append({

        "Order_ID":
            f"ORD{i:05d}",

        "Order_Date":
            order_date,

        "Customer_ID":
            selected_customer_id,

        "Product_ID":
            product_id,

        "Region":
            region,

        "Sales_Channel":
            sales_channel,

        "Quantity":
            quantity,

        "Unit_Price":
            unit_price,

        "Discount":
            discount,

        "Unit_Cost":
            unit_cost
    })


# ============================================================
# VERIFY ORDER GENERATION
# ============================================================

print(
    f"\nOrders generated in memory: {len(orders):,}"
)


# ------------------------------------------------------------
# Verify exact number of orders
# ------------------------------------------------------------

if len(orders) != NUM_ORDERS:

    raise ValueError(
        f"Order generation failed. "
        f"Expected {NUM_ORDERS:,} orders but generated "
        f"{len(orders):,}."
    )


# ------------------------------------------------------------
# Verify Order IDs
# ------------------------------------------------------------

generated_order_ids = [
    order["Order_ID"]
    for order in orders
]


unique_generated_order_ids = len(
    set(generated_order_ids)
)


if unique_generated_order_ids != NUM_ORDERS:

    raise ValueError(
        "Duplicate Order IDs detected during generation."
    )


# ------------------------------------------------------------
# Verify first and last Order ID
# ------------------------------------------------------------

expected_first_order_id = "ORD00001"

expected_last_order_id = (
    f"ORD{NUM_ORDERS:05d}"
)


if generated_order_ids[0] != expected_first_order_id:

    raise ValueError(
        f"Unexpected first Order ID: "
        f"{generated_order_ids[0]}"
    )


if generated_order_ids[-1] != expected_last_order_id:

    raise ValueError(
        f"Unexpected last Order ID: "
        f"{generated_order_ids[-1]}"
    )


print(
    "\nOrder ID sequence validation passed."
)


print(
    f"First Order ID: {generated_order_ids[0]}"
)


print(
    f"Last Order ID: {generated_order_ids[-1]}"
)


print(
    f"\nSuccessfully generated {len(orders):,} orders."
)


# ============================================================
# 18. CREATE ORDERS DATAFRAME
# ============================================================

orders_df = pd.DataFrame(
    orders
)


# ------------------------------------------------------------
# Verify DataFrame size
# ------------------------------------------------------------

print(
    "\nOrders DataFrame created."
)


print(
    "Orders shape:",
    orders_df.shape
)


print(
    "Rows:",
    len(orders_df)
)


print(
    "Columns:",
    len(orders_df.columns)
)


# ------------------------------------------------------------
# Final row-count validation
# ------------------------------------------------------------

if len(orders_df) != NUM_ORDERS:

    raise ValueError(
        f"Orders DataFrame contains "
        f"{len(orders_df):,} rows instead of "
        f"{NUM_ORDERS:,}."
    )


# ------------------------------------------------------------
# Final Order ID validation
# ------------------------------------------------------------

if orders_df["Order_ID"].nunique() != NUM_ORDERS:

    raise ValueError(
        "Duplicate Order IDs found in Orders DataFrame."
    )


# ------------------------------------------------------------
# Convert Order_Date to datetime
# ------------------------------------------------------------

orders_df["Order_Date"] = pd.to_datetime(
    orders_df["Order_Date"]
)


# ------------------------------------------------------------
# Validate order dates against project period
# ------------------------------------------------------------

invalid_start_dates = (
    orders_df["Order_Date"]
    < ORDER_START_DATE
).sum()


invalid_end_dates = (
    orders_df["Order_Date"]
    > ORDER_END_DATE
).sum()


print(
    "\nOrders before project start date:",
    invalid_start_dates
)


print(
    "Orders after project end date:",
    invalid_end_dates
)


if invalid_start_dates > 0:

    raise ValueError(
        "Orders found before project start date."
    )


if invalid_end_dates > 0:

    raise ValueError(
        "Orders found after project end date."
    )


print(
    "\nOrder date validation passed."
)


print(
    "\nFinancial metrics will now be calculated."
)
# ============================================================
# 19. FINANCIAL METRICS
# ============================================================

# Gross Sales
orders_df["Gross_Sales"] = (

    orders_df["Quantity"]

    *

    orders_df["Unit_Price"]

).round(2)


# Discount Amount
orders_df["Discount_Amount"] = (

    orders_df["Gross_Sales"]

    *

    orders_df["Discount"]

).round(2)


# Revenue
orders_df["Revenue"] = (

    orders_df["Gross_Sales"]

    -

    orders_df["Discount_Amount"]

).round(2)


# COGS
orders_df["COGS"] = (

    orders_df["Quantity"]

    *

    orders_df["Unit_Cost"]

).round(2)


# Gross Profit
orders_df["Gross_Profit"] = (

    orders_df["Revenue"]

    -

    orders_df["COGS"]

).round(2)


# Gross Margin
orders_df["Gross_Margin"] = (

    orders_df["Gross_Profit"]

    /

    orders_df["Revenue"]

)


print(
    "Financial metrics calculated successfully."
)


# ============================================================
# 20. ORDER GENERATION RESULTS
# ============================================================

print("\n" + "=" * 60)

print(
    "ORDER GENERATION RESULTS"
)

print("=" * 60)


print(
    "\nOrders generated:",
    len(orders_df)
)


print(
    "Number of columns:",
    len(orders_df.columns)
)


print(
    "\nOrders shape:",
    orders_df.shape
)


# ============================================================
# 21. ORDER ID VALIDATION
# ============================================================

unique_order_ids = (

    orders_df[
        "Order_ID"
    ]

    .nunique()
)


print(
    "\nUnique Order IDs:",
    unique_order_ids
)


if unique_order_ids != len(orders_df):

    raise ValueError(
        "Duplicate Order IDs found."
    )


# ============================================================
# 22. CUSTOMER ID VALIDATION
# ============================================================

invalid_customers = (

    set(
        orders_df[
            "Customer_ID"
        ]
    )

    -

    set(
        customers_df[
            "Customer_ID"
        ]
    )
)


print(
    "Invalid Customer IDs:",
    len(invalid_customers)
)


if invalid_customers:

    raise ValueError(
        "Invalid Customer IDs found."
    )


# ============================================================
# 23. PRODUCT ID VALIDATION
# ============================================================

invalid_products = (

    set(
        orders_df[
            "Product_ID"
        ]
    )

    -

    set(
        products_df[
            "Product_ID"
        ]
    )
)


print(
    "Invalid Product IDs:",
    len(invalid_products)
)


if invalid_products:

    raise ValueError(
        "Invalid Product IDs found."
    )


# ============================================================
# 24. MISSING VALUE VALIDATION
# ============================================================

missing_values = (

    orders_df

    .isna()

    .sum()

    .sum()
)


print(
    "Total missing values:",
    missing_values
)


if missing_values > 0:

    print(
        "\nMissing values by column:"
    )

    print(
        orders_df.isna().sum()
    )

    raise ValueError(
        "Missing values found."
    )


# ============================================================
# 25. QUANTITY VALIDATION
# ============================================================

print(
    "\nQuantity range:",

    orders_df[
        "Quantity"
    ].min(),

    "to",

    orders_df[
        "Quantity"
    ].max()
)


if (

    orders_df[
        "Quantity"
    ]

    <=

    0

).any():

    raise ValueError(
        "Invalid quantity found."
    )


# ============================================================
# 26. PRICE VALIDATION
# ============================================================

print(
    "Unit price range:",

    orders_df[
        "Unit_Price"
    ].min(),

    "to",

    orders_df[
        "Unit_Price"
    ].max()
)


if (

    orders_df[
        "Unit_Price"
    ]

    <=

    0

).any():

    raise ValueError(
        "Invalid unit price found."
    )


# ============================================================
# 27. COST VALIDATION
# ============================================================

print(
    "Unit cost range:",

    orders_df[
        "Unit_Cost"
    ].min(),

    "to",

    orders_df[
        "Unit_Cost"
    ].max()
)


if (

    orders_df[
        "Unit_Cost"
    ]

    <=

    0

).any():

    raise ValueError(
        "Invalid unit cost found."
    )


# ============================================================
# 28. DISCOUNT VALIDATION
# ============================================================

print(
    "\nDiscount values:"
)


print(

    sorted(

        orders_df[
            "Discount"
        ]

        .unique()
    )
)


if (

    (

        orders_df[
            "Discount"
        ]

        <

        0

    )

    |

    (

        orders_df[
            "Discount"
        ]

        >

        1

    )

).any():

    raise ValueError(
        "Invalid discount found."
    )


# ============================================================
# 29. FINANCIAL VALIDATION
# ============================================================

print("\n" + "=" * 60)

print(
    "FINANCIAL VALIDATION"
)

print("=" * 60)


# Revenue
revenue_check = (

    orders_df[
        "Revenue"
    ]

    -

    (

        orders_df[
            "Gross_Sales"
        ]

        -

        orders_df[
            "Discount_Amount"
        ]

    )

).abs().max()


print(
    "\nMaximum Revenue calculation difference:",
    revenue_check
)


# COGS
cogs_check = (

    orders_df[
        "COGS"
    ]

    -

    (

        orders_df[
            "Quantity"
        ]

        *

        orders_df[
            "Unit_Cost"
        ]

    )

).abs().max()


print(
    "Maximum COGS calculation difference:",
    cogs_check
)


# Gross Profit
profit_check = (

    orders_df[
        "Gross_Profit"
    ]

    -

    (

        orders_df[
            "Revenue"
        ]

        -

        orders_df[
            "COGS"
        ]

    )

).abs().max()


print(
    "Maximum Profit calculation difference:",
    profit_check
)


# Gross Margin
margin_check = (

    orders_df[
        "Gross_Margin"
    ]

    -

    (

        orders_df[
            "Gross_Profit"
        ]

        /

        orders_df[
            "Revenue"
        ]

    )

).abs().max()


print(
    "Maximum Margin calculation difference:",
    margin_check
)


# Non-positive revenue
negative_revenue = (

    orders_df[
        "Revenue"
    ]

    <=

    0

).sum()


print(
    "Orders with non-positive revenue:",
    negative_revenue
)


# Negative profit
negative_profit = (

    orders_df[
        "Gross_Profit"
    ]

    <

    0

).sum()


print(
    "Orders with negative gross profit:",
    negative_profit
)


# ============================================================
# 30. FINANCIAL VALIDATION RULES
# ============================================================

if revenue_check > 0.01:

    raise ValueError(
        "Revenue calculation failed."
    )


if cogs_check > 0.01:

    raise ValueError(
        "COGS calculation failed."
    )


if profit_check > 0.01:

    raise ValueError(
        "Profit calculation failed."
    )


if negative_revenue > 0:

    raise ValueError(
        "Non-positive revenue found."
    )


if negative_profit > 0:

    raise ValueError(
        "Negative gross profit found."
    )


print(
    "\nFinancial validation passed."
)


# ============================================================
# 31. DISTRIBUTION CHECKS
# ============================================================

print("\n" + "=" * 60)

print(
    "DISTRIBUTION CHECKS"
)

print("=" * 60)


# ------------------------------------------------------------
# Customer Segment
# ------------------------------------------------------------

segment_distribution = (

    orders_df

    .merge(

        customers_df[
            [
                "Customer_ID",
                "Customer_Segment"
            ]
        ],

        on="Customer_ID",

        how="left"
    )

    ["Customer_Segment"]

    .value_counts()
)


print(
    "\nOrders by customer segment:"
)


print(
    segment_distribution
)


# ------------------------------------------------------------
# Region
# ------------------------------------------------------------

print(
    "\nOrders by region:"
)


print(

    orders_df[
        "Region"
    ]

    .value_counts()
)


# ------------------------------------------------------------
# Sales Channel
# ------------------------------------------------------------

print(
    "\nOrders by sales channel:"
)


print(

    orders_df[
        "Sales_Channel"
    ]

    .value_counts()
)


# ------------------------------------------------------------
# Product Category
# ------------------------------------------------------------

category_distribution = (

    orders_df

    .merge(

        products_df[
            [
                "Product_ID",
                "Category"
            ]
        ],

        on="Product_ID",

        how="left"
    )

    ["Category"]

    .value_counts()
)


print(
    "\nOrders by product category:"
)


print(
    category_distribution
)


# ============================================================
# 32. FINANCIAL SUMMARY
# ============================================================

print("\n" + "=" * 60)

print(
    "FINANCIAL SUMMARY"
)

print("=" * 60)


total_gross_sales = (

    orders_df[
        "Gross_Sales"
    ]

    .sum()
)


total_discount = (

    orders_df[
        "Discount_Amount"
    ]

    .sum()
)


total_revenue = (

    orders_df[
        "Revenue"
    ]

    .sum()
)


total_cogs = (

    orders_df[
        "COGS"
    ]

    .sum()
)


total_profit = (

    orders_df[
        "Gross_Profit"
    ]

    .sum()
)


overall_margin = (

    total_profit

    /

    total_revenue
)


print(
    f"\nGross Sales: ₹{total_gross_sales:,.2f}"
)


print(
    f"Discount Amount: ₹{total_discount:,.2f}"
)


print(
    f"Revenue: ₹{total_revenue:,.2f}"
)


print(
    f"COGS: ₹{total_cogs:,.2f}"
)


print(
    f"Gross Profit: ₹{total_profit:,.2f}"
)


print(
    f"Gross Margin: {overall_margin:.2%}"
)


# ============================================================
# 33. FIRST 10 ORDERS
# ============================================================

print("\n" + "=" * 60)

print(
    "FIRST 10 ORDERS"
)

print("=" * 60)


print(

    orders_df[
        [
            "Order_ID",
            "Order_Date",
            "Customer_ID",
            "Product_ID",
            "Region",
            "Sales_Channel",
            "Quantity",
            "Unit_Price",
            "Discount",
            "Unit_Cost",
            "Gross_Sales",
            "Discount_Amount",
            "Revenue",
            "COGS",
            "Gross_Profit",
            "Gross_Margin"
        ]
    ]

    .head(10)

    .to_string(
        index=False
    )
)


# ============================================================
# 34. FINAL STATUS
# ============================================================

print("\n" + "=" * 60)

print(
    "VALIDATION COMPLETE"
)

print("=" * 60)


print(
    f"\n{NUM_ORDERS:,} orders generated and validated successfully."
)


print(
    "\nThe Excel workbook has NOT been modified."
)


print(
    "\nNext step: write the validated order data to the Orders sheet."
)
# ============================================================
# 35. SAVE VALIDATED ORDERS FOR WRITE SCRIPT
# ============================================================

TEMP_ORDERS_FILE = (
    PROJECT_DIR
    / "03_Python"
    / "_validated_orders.csv"
)

orders_df.to_csv(
    TEMP_ORDERS_FILE,
    index=False
)


print(
    "\nValidated orders temporarily saved to:"
)

print(
    TEMP_ORDERS_FILE
)


# ============================================================
# 36. RETURN VALIDATED ORDERS
# ============================================================

GENERATED_ORDERS = orders_df.copy()